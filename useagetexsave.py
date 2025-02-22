from calendar import isleap
from datetime import datetime, date
import time
from pathlib import Path
from openpyxl import Workbook, load_workbook
import os

class AgeCalculator:
def __init__(self):
self.months = ['January', 'February', 'March', 'April', 'May', 'June', 
'July', 'August', 'September', 'October', 'November', 'December']
self.excel_file = 'age_calculations.xlsx'

def judge_leap_year(self, year):
return isleap(year)

def month_days(self, month, leap_year):
if month in [1, 3, 5, 7, 8, 10, 12]:
return 31
elif month in [4, 6, 9, 11]:
return 30
elif month == 2 and leap_year:
return 29
elif month == 2 and (not leap_year):
return 28
return 0

def validate_name(self, name):
if not name.strip():
raise ValueError("Name cannot be empty")
if len(name) > 50:
raise ValueError("Name is too long (maximum 50 characters)")
if any(char.isdigit() for char in name):
raise ValueError("Name should not contain numbers")
return name.strip()

def get_valid_date(self):
while True:
try:
print("\nEnter your date of birth:")
year = int(input("Year (1900-present): "))
month = int(input("Month (1-12): "))
day = int(input("Day (1-31): "))
current_year = datetime.now().year
if not (1900 <= year <= current_year):
print(f"Invalid year. Please enter a year between 1900 and {current_year}.")
continue
if not (1 <= month <= 12):
print("Invalid month. Please enter a month between 1 and 12.")
continue
max_days = self.month_days(month, self.judge_leap_year(year))
if not (1 <= day <= max_days):
print(f"Invalid day. For {self.months[month-1]} {year}, please enter a day between 1 and {max_days}.")
continue
return year, month, day
except ValueError:
print("Please enter valid numbers.")

def calculate_age(self):
try:
print("Welcome to the Age Calculator!")
name = input("Enter your name: ")
name = self.validate_name(name)
birth_year, birth_month, birth_day = self.get_valid_date()
today = date.today()
age = today.year - birth_year
if today.month < birth_month or (today.month == birth_month and today.day < birth_day):
age -= 1
total_months = age * 12 + (today.month - birth_month)
if today.day < birth_day:
total_months -= 1
start_date = date(birth_year, birth_month, birth_day)
days = (today - start_date).days
next_milestone = (age // 10 + 1) * 10
days_to_milestone = (date(birth_year + next_milestone, birth_month, birth_day) - today).days
# Create results dictionary for Excel
results = {
'Name': name,
'Birth Date': f"{birth_year}-{birth_month:02d}-{birth_day:02d}",
'Age (Years)': age,
'Age (Months)': total_months,
'Age (Days)': days,
'Next Milestone': next_milestone,
'Days to Milestone': days_to_milestone,
'Sunrises': age * 365,
'Sleep Days': int(days * 8/24),
'Heartbeats': int(days * 24 * 60 * 80),
'Calculation Date': today.strftime("%Y-%m-%d")
}
# Display results
print("\nResults:")
for key, value in results.items():
print(f"{key}: {value}")
return results
except Exception as e:
print(f"An error occurred: {str(e)}")
return None

def save_to_excel(self, results):
try:
if not results:
print("\nNo results to save.")
return

# Create or load workbook
if os.path.exists(self.excel_file):
wb = load_workbook(self.excel_file)
ws = wb.active
else:
wb = Workbook()
ws = wb.active
ws.title = "Age Calculations"
# Add headers
headers = list(results.keys())
for col, header in enumerate(headers, 1):
ws.cell(row=1, column=col, value=header)

# Find next empty row
next_row = ws.max_row + 1
if next_row == 2 and ws.cell(row=1, column=1).value is None:
next_row = 1

# Add data
for col, value in enumerate(results.values(), 1):
ws.cell(row=next_row, column=col, value=value)

# Adjust column widths
for column in ws.columns:
max_length = 0
column_letter = column[0].column_letter
for cell in column:
try:
if len(str(cell.value)) > max_length:
max_length = len(str(cell.value))
except:
pass
adjusted_width = (max_length + 2)
ws.column_dimensions[column_letter].width = adjusted_width

wb.save(self.excel_file)
print(f"\nResults saved to {self.excel_file}")
except Exception as e:
print(f"\nError saving to Excel: {str(e)}")

def load_from_excel(self):
try:
if not os.path.exists(self.excel_file):
print("\nNo saved calculations found.")
return

wb = load_workbook(self.excel_file)
ws = wb.active
print("\nPrevious Calculations:")
print("-" * 80)
headers = [cell.value for cell in ws[1]]
for row in ws.iter_rows(min_row=2):
values = [cell.value for cell in row]
if any(values): # Skip empty rows
print("\nCalculation Record:")
for header, value in zip(headers, values):
print(f"{header}: {value}")
print("-" * 40)
except Exception as e:
print(f"\nError loading from Excel: {str(e)}")

def main():
calculator = AgeCalculator()
while True:
try:
print("\n=== Age Calculator Menu ===")
print("1. Calculate Age")
print("2. View Previous Calculations")
print("3. Exit")
choice = input("\nEnter your choice (1-3): ").strip()
if choice == '1':
results = calculator.calculate_age()
if results:
save = input("\nWould you like to save these results? (y/n): ").strip().lower()
if save == 'y':
calculator.save_to_excel(results)
elif choice == '2':
calculator.load_from_excel()
elif choice == '3':
print("\nThank you for using the Age Calculator!")
break
else:
print("\nInvalid choice. Please enter 1, 2, or 3.")
except KeyboardInterrupt:
print("\n\nProgram interrupted by user. Exiting...")
break
except Exception as e:
print(f"\nAn unexpected error occurred: {str(e)}")

if __name__ == "__main__":
main()